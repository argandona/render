from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction
from .models import SST, Suministro, Distrito, EstadoSST, EstadoSuministro
import pandas as pd
from decimal import Decimal
from django.db import models as django_models
from django.core.cache import cache

from django.http import JsonResponse
from django.views.decorators.http import require_POST
import json

from django.core.paginator import Paginator


from django.views.decorators.csrf import csrf_exempt
from datetime import datetime


from django.core.paginator import Paginator
from django.core.cache import cache
from django.db.models import Q, Count

import pandas as pd
from django.http import HttpResponse
from io import BytesIO
from django.db.models import Q



@login_required
def dashboard(request):
    return render(request, 'gestion/dashboard.html')

@login_required
def sst_list(request):
    ssts = SST.objects.select_related('distrito', 'estado_sst').all()
    distritos = Distrito.objects.all()
    
    context = {
        'ssts': ssts,
        'distritos': distritos,
    }
    return render(request, 'gestion/sst.html', context)


@login_required
def cargar_excel_suministros(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file = request.FILES['excel_file']
        
        try:
            # Leer el Excel
            df = pd.read_excel(excel_file)
            
            # Limpiar nombres de columnas
            df.columns = df.columns.str.strip()
            
            # Mapeo de columnas del Excel a nuestro modelo
            columnas_map = {
                'ITEM': 'Item',
                'N° OT': 'No. OT',
                'SUMINISTRO': 'Suministro',
                'DIRECCIÓN': 'Dirección',
                'DISTRITO': 'Distrito',
                'MEDIDOR': 'Medidor',
                'MARCA MEDIDOR': 'Marca',
                'FASE': 'Fase',
                'POTENCIA': 'Potencia',
                'FECHA_PRIMER_ENVIO': 'Fecha Primer Envío',
                'FECHA_PROGRAMADA': 'Fec. Prog.',
                'HORA_INICIO_PROG': 'Hor. Prog.',
                'HORA_FIN_PROG': 'Hor. Fin. Prog.',
                'FECHA_EJECUCION': 'Fecha Ejecución',
                'EJECUTADO_POR': 'Ejecutado Por',
                'CONTACTO': 'Contacto',
                'TELEFONO': 'Teléfono',
                'LATITUD': 'Latitud',
                'LONGITUD': 'Longitud',
                'OBSERVACION_CONTRATISTA': 'Observación Contratista',
                'ESTADO': 'Estado',
            }
            
            # Verificar columnas críticas
            columnas_criticas = ['Item', 'No. OT', 'Suministro', 'Dirección', 'Distrito']
            columnas_faltantes = [col for col in columnas_criticas if col not in df.columns]
            
            if columnas_faltantes:
                messages.error(request, f'❌ Faltan columnas: {", ".join(columnas_faltantes)}')
                return redirect('sst_list')
            
            # Obtener o crear estados por defecto
            estado_sst_default, _ = EstadoSST.objects.get_or_create(
                estado='PENDIENTE',
                defaults={'descripcion': 'Estado por defecto', 'color': '#FFA500'}
            )
            
            estado_suministro_default, _ = EstadoSuministro.objects.get_or_create(
                estado_suministro='PENDIENTE',
                defaults={'descripcion': 'Estado por defecto', 'color': '#FFA500'}
            )
            
            # Contadores
            ssts_creadas = 0
            suministros_creados = 0
            suministros_actualizados = 0
            errores = []
            
            # Agrupar por N° OT
            ots_unicas = df['No. OT'].dropna().unique()
            
            with transaction.atomic():
                ssts_procesadas = set()  # Para trackear SSTs que necesitan actualización
                
                for ot in ots_unicas:
                    try:
                        # Filtrar suministros de esta OT
                        suministros_ot = df[df['No. OT'] == ot]
                        
                        if len(suministros_ot) == 0:
                            continue
                        
                        # Extraer código SST del N° OT
                        codigo_sst = str(ot).strip()[:7] if len(str(ot).strip()) >= 7 else str(ot).strip()
                        
                        # Obtener distrito más frecuente
                        distrito_principal_nombre = suministros_ot['Distrito'].mode()[0] if not suministros_ot['Distrito'].mode().empty else suministros_ot['Distrito'].iloc[0]
                        distrito_principal, _ = Distrito.objects.get_or_create(nombre_distrito=str(distrito_principal_nombre).strip())
                        
                        # Primera dirección
                        direccion_principal = str(suministros_ot['Dirección'].iloc[0]).strip() if pd.notna(suministros_ot['Dirección'].iloc[0]) else 'Sin dirección'
                        
                        # Crear o actualizar SST
                        sst, created = SST.objects.get_or_create(
                            sst=codigo_sst,
                            defaults={
                                'direccion': direccion_principal,
                                'distrito': distrito_principal,
                                'estado_sst': estado_sst_default,
                                'monto_proyectado': Decimal('0.00')
                            }
                        )
                        
                        if created:
                            ssts_creadas += 1
                        
                        ssts_procesadas.add(sst.id)  # Guardar ID de SST procesada
                        
                        # Crear cada suministro
                        for index, row in suministros_ot.iterrows():
                            try:
                                # Extraer datos
                                item = int(row['Item']) if pd.notna(row['Item']) else 0
                                no_ot = str(row['No. OT']).strip() if pd.notna(row['No. OT']) else ''
                                suministro_codigo = str(row['Suministro']).strip() if pd.notna(row['Suministro']) else ''
                                medidor = str(row['Medidor']).strip() if pd.notna(row['Medidor']) else None
                                marca_medidor = str(row['Marca']).strip() if pd.notna(row['Marca']) else None
                                fase = str(row['Fase']).strip() if pd.notna(row['Fase']) else None
                                potencia = str(row['Potencia']).strip() if pd.notna(row['Potencia']) else None
                                direccion = str(row['Dirección']).strip() if pd.notna(row['Dirección']) else ''
                                distrito_nombre = str(row['Distrito']).strip() if pd.notna(row['Distrito']) else ''
                                
                                # Datos adicionales
                                latitud = row.get('Latitud')
                                longitud = row.get('Longitud')
                                contacto = str(row.get('Contacto', '')).strip() if pd.notna(row.get('Contacto')) else None
                                telefono = str(row.get('Teléfono', '')).strip() if pd.notna(row.get('Teléfono')) else None
                                ejecutado_por = str(row.get('Ejecutado Por', '')).strip() if pd.notna(row.get('Ejecutado Por')) else None
                                observacion_contratista = str(row.get('Observación Contratista', '')).strip() if pd.notna(row.get('Observación Contratista')) else None
                                
                                # Fechas y horas
                                fecha_primer_envio = pd.to_datetime(row.get('Fecha Primer Envío'), errors='coerce') if pd.notna(row.get('Fecha Primer Envío')) else None
                                fecha_programada = pd.to_datetime(row.get('Fec. Prog.'), errors='coerce') if pd.notna(row.get('Fec. Prog.')) else None
                                fecha_ejecucion = pd.to_datetime(row.get('Fecha Ejecución'), errors='coerce') if pd.notna(row.get('Fecha Ejecución')) else None
                                
                                # Convertir a TimeField si es posible
                                hora_inicio_prog = None
                                hora_fin_prog = None
                                if pd.notna(row.get('Hor. Prog.')):
                                    try:
                                        hora_inicio_prog = pd.to_datetime(str(row['Hor. Prog.']), format='%H:%M:%S', errors='coerce').time()
                                    except:
                                        pass
                                
                                if pd.notna(row.get('Hor. Fin. Prog.')):
                                    try:
                                        hora_fin_prog = pd.to_datetime(str(row['Hor. Fin. Prog.']), format='%H:%M:%S', errors='coerce').time()
                                    except:
                                        pass
                                
                                # Obtener o crear distrito del suministro
                                distrito_suministro, _ = Distrito.objects.get_or_create(nombre_distrito=distrito_nombre)
                                
                                # Obtener o crear estado del suministro
                                estado_nombre = str(row.get('Estado', '')).strip() if pd.notna(row.get('Estado')) else ''
                                if estado_nombre:
                                    estado_sum, _ = EstadoSuministro.objects.get_or_create(
                                        estado_suministro=estado_nombre,
                                        defaults={'descripcion': '', 'color': '#007bff'}
                                    )
                                else:
                                    estado_sum = estado_suministro_default
                                
                                # Verificar si existe
                                suministro_existente = Suministro.objects.filter(
                                    sst=sst,
                                    suministro=suministro_codigo
                                ).first()
                                
                                if suministro_existente:
                                    # Actualizar
                                    suministro_existente.item = item
                                    suministro_existente.no_ot = no_ot
                                    suministro_existente.medidor = medidor
                                    suministro_existente.marca_medidor = marca_medidor
                                    suministro_existente.fase = fase
                                    suministro_existente.potencia = potencia
                                    suministro_existente.direccion = direccion
                                    suministro_existente.distrito = distrito_suministro
                                    suministro_existente.latitud = latitud
                                    suministro_existente.longitud = longitud
                                    suministro_existente.contacto = contacto
                                    suministro_existente.telefono = telefono
                                    suministro_existente.fecha_primer_envio = fecha_primer_envio
                                    suministro_existente.fecha_programada = fecha_programada
                                    suministro_existente.hora_inicio_programada = hora_inicio_prog
                                    suministro_existente.hora_fin_programada = hora_fin_prog
                                    suministro_existente.fecha_ejecucion = fecha_ejecucion
                                    suministro_existente.ejecutado_por = ejecutado_por
                                    suministro_existente.estado_suministro = estado_sum
                                    suministro_existente.observacion_contratista = observacion_contratista
                                    suministro_existente.save()
                                    suministros_actualizados += 1
                                else:
                                    # Crear nuevo
                                    Suministro.objects.create(
                                        sst=sst,
                                        item=item,
                                        no_ot=no_ot,
                                        suministro=suministro_codigo,
                                        medidor=medidor,
                                        marca_medidor=marca_medidor,
                                        fase=fase,
                                        potencia=potencia,
                                        direccion=direccion,
                                        distrito=distrito_suministro,
                                        latitud=latitud,
                                        longitud=longitud,
                                        contacto=contacto,
                                        telefono=telefono,
                                        fecha_primer_envio=fecha_primer_envio,
                                        fecha_programada=fecha_programada,
                                        hora_inicio_programada=hora_inicio_prog,
                                        hora_fin_programada=hora_fin_prog,
                                        fecha_ejecucion=fecha_ejecucion,
                                        ejecutado_por=ejecutado_por,
                                        estado_suministro=estado_sum,
                                        observacion_contratista=observacion_contratista
                                    )
                                    suministros_creados += 1
                                
                            except Exception as e:
                                errores.append(f"Fila {index + 2}: {str(e)}")
                                continue
                        
                    except Exception as e:
                        errores.append(f"OT {ot}: {str(e)}")
                        continue
                
                # ⭐ ACTUALIZAR ESTADOS DE TODAS LAS SST PROCESADAS
                for sst_id in ssts_procesadas:
                    try:
                        sst = SST.objects.get(id=sst_id)
                        sst.actualizar_estado_segun_suministros()
                    except Exception as e:
                        errores.append(f"Error actualizando SST {sst_id}: {str(e)}")
            
            # Mensajes
            mensaje_exito = []
            if ssts_creadas > 0:
                mensaje_exito.append(f'{ssts_creadas} SST creadas')
            if suministros_creados > 0:
                mensaje_exito.append(f'{suministros_creados} suministros creados')
            if suministros_actualizados > 0:
                mensaje_exito.append(f'{suministros_actualizados} suministros actualizados')
            
            if mensaje_exito:
                messages.success(request, f'✅ {", ".join(mensaje_exito)}.')
            
            if errores:
                messages.warning(request, f'⚠️ {len(errores)} errores encontrados.')
                for error in errores[:5]:
                    messages.warning(request, error)
            
            return redirect('sst_list')
            
        except Exception as e:
            messages.error(request, f'❌ Error: {str(e)}')
            return redirect('sst_list')
    
    return redirect('sst_list')




@login_required
def suministro_list(request):

    # Filtros
    sst_filter = request.GET.get('sst', '')
    distrito_filter = request.GET.get('distrito', '')
    estado_filter = request.GET.get('estado', '')
    search = request.GET.get('search', '')

    todas_sst = SST.objects.select_related('distrito', 'estado_sst').order_by('sst')

    # 🔹 Query base (SIEMPRE primero)
    suministros = Suministro.objects.select_related(
        'sst', 'distrito', 'estado_suministro'
    ).all()

    # 🔹 Filtros
    if sst_filter:
        suministros = suministros.filter(sst__sst__icontains=sst_filter)

    if distrito_filter:
        suministros = suministros.filter(distrito__nombre_distrito=distrito_filter)

    if estado_filter:
        suministros = suministros.filter(
            estado_suministro__estado_suministro=estado_filter
        )

    if search:
        suministros = suministros.filter(
            Q(suministro__icontains=search) |
            Q(medidor__icontains=search) |
            Q(direccion__icontains=search)
        )

    # 🔹 Paginación (DESPUÉS de filtros)
    paginator = Paginator(suministros, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # 🔹 Cache distritos (bien)
    distritos = cache.get('distritos')
    if not distritos:
        distritos = Distrito.objects.all()
        cache.set('distritos', distritos, 3600)

    estados = EstadoSuministro.objects.all()
    ssts = SST.objects.all()

    # 🔹 Estadísticas SST
    sst_stats = SST.objects.values('estado_sst__estado').annotate(
        cantidad=Count('id')
    )
    sst_por_estado = {i['estado_sst__estado']: i['cantidad'] for i in sst_stats}

    # 🔹 Estadísticas Suministros
    suministro_stats = Suministro.objects.values(
        'estado_suministro__estado_suministro'
    ).annotate(cantidad=Count('id'))
    suministros_por_estado = {
        i['estado_suministro__estado_suministro']: i['cantidad']
        for i in suministro_stats
    }

    context = {
        'suministros': page_obj,  # 👈 usa el paginado
        'page_obj': page_obj,
        'paginator': paginator,

        'distritos': distritos,
        'estados': estados,
        'ssts': ssts,
        'todas_sst': todas_sst,

        'sst_filter': sst_filter,
        'distrito_filter': distrito_filter,
        'estado_filter': estado_filter,
        'search': search,

        # Estadísticas SST
        'sst_ejecutado': sst_por_estado.get('Ejecutado', 0),
        'sst_admisible': sst_por_estado.get('Admisible', 0),
        'sst_en_ejecucion': sst_por_estado.get('En Ejecución', 0),
        'sst_pendiente': sst_por_estado.get('Pendiente', 0),

        # Estadísticas Suministros
        'sum_asignado': suministros_por_estado.get('ASIGNADO', 0)
                        + suministros_por_estado.get('Asignado', 0),
        'sum_devuelto': suministros_por_estado.get('DEVUELTO', 0)
                        + suministros_por_estado.get('Devuelto', 0),
        'sum_ejecutado': suministros_por_estado.get('EJECUTADO', 0)
                        + suministros_por_estado.get('Ejecutado', 0),

        'total_suministros': suministros.count(),
    }

    return render(request, 'gestion/suministros.html', context)




@login_required
@require_POST
def actualizar_suministro(request, suministro_id):
    try:
        suministro = Suministro.objects.get(id=suministro_id)
        
        # Obtener datos del POST
        data = json.loads(request.body)
        # Guardar el monto anterior para comparar
        monto_anterior = suministro.monto
        
        # Actualizar estado
        estado_id = data.get('estado_suministro')
        if estado_id:
            try:
                estado = EstadoSuministro.objects.get(id=estado_id)
                suministro.estado_suministro = estado
            except EstadoSuministro.DoesNotExist:
                pass
        
        # Actualizar fecha de ejecución
        fecha_ejecucion = data.get('fecha_ejecucion')
        if fecha_ejecucion:
            from datetime import datetime
            suministro.fecha_ejecucion = datetime.strptime(fecha_ejecucion, '%Y-%m-%d').date()
        
        # Actualizar ejecutado por
        ejecutado_por = data.get('ejecutado_por')
        if ejecutado_por:
            suministro.ejecutado_por = ejecutado_por
            
        # ✅ ACTUALIZAR MONTO (IMPORTANTE)
        monto = data.get('monto')
        if monto is not None:
            try:
                suministro.monto = Decimal(str(monto))
            except (ValueError, TypeError):
                pass
        
        observacion = data.get('observacion_contratista')
        if observacion is not None:
            suministro.observacion_contratista = observacion
        
        suministro.save()
        
        # ✅ ACTUALIZAR MONTO DE LA SST SI CAMBIÓ EL MONTO
        if monto_anterior != suministro.monto:
            suministro.sst.actualizar_monto_total()
        
        # Actualizar estado de la SST
        suministro.sst.actualizar_estado_segun_suministros()
        
        return JsonResponse({
            'success': True,
            'message': 'Suministro actualizado correctamente'
        })
        
    except Suministro.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Suministro no encontrado'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=400)
        
           
        
        # Actualizar observaciones
        observacion = data.get('observacion_contratista')
        if observacion is not None:
            suministro.observacion_contratista = observacion
        
        suministro.save()
        
        # ⭐ ACTUALIZAR ESTADO DE LA SST AUTOMÁTICAMENTE
        suministro.sst.actualizar_estado_segun_suministros()
        
        return JsonResponse({
            'success': True,
            'message': 'Suministro actualizado correctamente'
        })
        
    except Suministro.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Suministro no encontrado'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=400)
        
        
@login_required
@csrf_exempt
def agregar_suministro_adicional(request):
    """Vista para agregar suministros adicionales vía AJAX"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            # Validar datos requeridos
            sst_id = data.get('sst_id')
            if not sst_id:
                return JsonResponse({
                    'success': False,
                    'message': 'Debe seleccionar una SST'
                }, status=400)
            
            motivo_adicional = data.get('motivo_adicional', '').strip()
            if not motivo_adicional:
                return JsonResponse({
                    'success': False,
                    'message': 'El motivo del suministro adicional es obligatorio'
                }, status=400)
            
            # Obtener la SST
            try:
                sst = SST.objects.get(id=sst_id)
            except SST.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'message': 'SST no encontrada'
                }, status=404)
            
            # Generar número de suministro automático si está vacío
            suministro_num = data.get('suministro', '').strip()
            if not suministro_num:
                # Contar suministros adicionales existentes en esta SST
                count_adicionales = Suministro.objects.filter(
                    sst=sst, 
                    tipo_suministro='ADICIONAL'
                ).count()
                suministro_num = f"{sst.sst}-AD{count_adicionales + 1:03d}"
            
            # Generar item automático
            ultimo_suministro = Suministro.objects.filter(sst=sst).order_by('-item').first()
            nuevo_item = ultimo_suministro.item + 1 if ultimo_suministro else 1
            
            # Obtener estado del suministro
            estado_sum = None
            estado_id = data.get('estado_suministro')
            if estado_id:
                try:
                    estado_sum = EstadoSuministro.objects.get(id=estado_id)
                except EstadoSuministro.DoesNotExist:
                    pass
            
            # Si no se especifica estado, usar por defecto
            if not estado_sum:
                estado_sum, _ = EstadoSuministro.objects.get_or_create(
                    estado_suministro='PENDIENTE',
                    defaults={'descripcion': 'Pendiente de ejecución', 'color': '#FFA500'}
                )
                
                
            # ✅ OBTENER MONTO (si no viene, se calculará automáticamente)
            monto = Decimal('0.00')
            monto_data = data.get('monto')
            if monto_data:
                try:
                    monto = Decimal(str(monto_data))
                except:
                    pass    
            
            # Crear el suministro adicional
            suministro = Suministro(
                sst=sst,
                tipo_suministro='ADICIONAL',
                item=nuevo_item,
                no_ot=sst.sst,  # Usar SST como N° OT
                suministro=suministro_num,
                monto=monto,  # ✅ AGREGAR MONTO
                estado_suministro=estado_sum,
                direccion=sst.direccion,
                distrito=sst.distrito,
                motivo_adicional=motivo_adicional,
                fecha_ejecucion=data.get('fecha_ejecucion') or None,
                ejecutado_por=data.get('ejecutado_por', '').strip() or None,
                contacto=data.get('contacto', '').strip() or None,
                telefono=data.get('telefono', '').strip() or None,
                solicitado_por=data.get('solicitado_por', '').strip() or None,
                observacion_contratista=data.get('observacion_contratista', '').strip() or None,
                fecha_identificacion=datetime.now().date()  # Fecha actual
            )
            
            suministro.save()
            
            # Actualizar estado de la SST
            sst.actualizar_estado_segun_suministros()
            
            return JsonResponse({
                'success': True,
                'message': f'Suministro adicional "{suministro_num}" creado exitosamente',
                'suministro_id': suministro.id,
                'suministro_numero': suministro_num
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'message': f'Error: {str(e)}'
            }, status=500)
    
    return JsonResponse({
        'success': False,
        'message': 'Método no permitido'
    }, status=405)


@login_required
def obtener_info_sst(request, sst_id):
    """Obtener información de una SST específica para el modal"""
    try:
        sst = SST.objects.select_related('distrito', 'estado_sst').get(id=sst_id)
        
        # Obtener estadísticas de la SST
        total_suministros = sst.suministros.count()
        suministros_adicionales = sst.suministros.filter(tipo_suministro='ADICIONAL').count()
        
        # Generar sugerencia de número de suministro
        ultimo_adicional = sst.suministros.filter(tipo_suministro='ADICIONAL').order_by('-suministro').first()
        if ultimo_adicional:
            try:
                # Intentar extraer número del último suministro adicional
                ultimo_num = ultimo_adicional.suministro.split('-')[-1]
                if ultimo_num.startswith('AD'):
                    num = int(ultimo_num[2:]) + 1
                    sugerencia = f"{sst.sst}-AD{num:03d}"
                else:
                    sugerencia = f"{sst.sst}-AD001"
            except:
                sugerencia = f"{sst.sst}-AD001"
        else:
            sugerencia = f"{sst.sst}-AD001"
        
        data = {
            'sst': sst.sst,
            'direccion': sst.direccion,
            'distrito': sst.distrito.nombre_distrito if sst.distrito else 'Sin distrito',
            'estado': sst.estado_sst.estado if sst.estado_sst else 'Sin estado',
            'total_suministros': total_suministros,
            'suministros_adicionales': suministros_adicionales,
            'sugerencia_suministro': sugerencia
        }
        
        return JsonResponse({
            'success': True,
            'data': data
        })
        
    except SST.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'SST no encontrada'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=500)   
        
        
def suministros_view(request):
    context = {
        # SST
        "sst_ejecutadas": SST.objects.filter(estado_sst__estado="Ejecutada").count(),
        "sst_admisibles": SST.objects.filter(estado_sst__estado="Admisible").count(),
        "sst_en_ejecucion": SST.objects.filter(estado_sst__estado="En Ejecución").count(),

        # SUMINISTROS
        "suministros_ejecutados": Suministro.objects.filter(
            estado_suministro__estado_suministro="Ejecutado"
        ).count(),
        "suministros_asignados": Suministro.objects.filter(
            estado_suministro__estado_suministro="Asignado"
        ).count(),
        "suministros_devueltos": Suministro.objects.filter(
            estado_suministro__estado_suministro="Devuelto"
        ).count(),

        # lo que ya tenías
        "suministros": Suministro.objects.select_related(
            "sst", "estado_suministro", "distrito"
        ),
    }

    return render(request, "gestion/suministros.html", context)             




@login_required
@require_POST
def eliminar_suministro(request, suministro_id):
    try:
        suministro = Suministro.objects.get(id=suministro_id)
        sst = suministro.sst  # Guardar referencia antes de eliminar
        
        # Eliminar suministro
        suministro.delete()
        
        # Actualizar estado de la SST automáticamente
        sst.actualizar_estado_segun_suministros()
        
        return JsonResponse({
            'success': True,
            'message': f'Suministro {suministro.suministro} eliminado correctamente'
        })
        
    except Suministro.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Suministro no encontrado'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': str(e)
        }, status=400)
        
 
 
@login_required
def buscar_sst(request):
    """Vista para buscar SST con autocompletado"""
    query = request.GET.get('q', '').strip()
    
    if len(query) < 2:
        return JsonResponse({
            'success': True,
            'data': []
        })
    
    try:
        # Buscar por código SST o dirección
        sst_list = SST.objects.filter(
            Q(sst__icontains=query) | 
            Q(direccion__icontains=query)
        ).select_related('distrito').order_by('sst')[:10]  # Limitar a 10 resultados
        
        results = []
        for sst in sst_list:
            # Asegúrate de que estos campos existan en tu modelo
            total_original = sst.suministros.filter(tipo_suministro='ORIGINAL').count()
            total_adicional = sst.suministros.filter(tipo_suministro='ADICIONAL').count()
            
            results.append({
                'id': sst.id,
                'sst': sst.sst,
                'direccion': sst.direccion,
                'distrito': sst.distrito.nombre_distrito if sst.distrito else '',
                'estado': sst.estado_sst.estado if sst.estado_sst else '',  # Corregido
                'total_suministros': f"{total_original}/{total_adicional}"
            })
        
        return JsonResponse({
            'success': True,
            'data': results
        })
    
    except Exception as e:
        # Registrar el error para debugging
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error en buscar_sst: {str(e)}")
        
        return JsonResponse({
            'success': False,
            'message': 'Error interno del servidor',
            'data': []
        }, status=500)
        
        
        

@login_required
def descargar_excel_suministros(request):
    """Descargar suministros filtrados en Excel"""
    try:
        # Obtener los mismos filtros que en suministro_list
        sst_filter = request.GET.get('sst', '')
        distrito_filter = request.GET.get('distrito', '')
        estado_filter = request.GET.get('estado', '')
        search = request.GET.get('search', '')
        
        # Aplicar los mismos filtros que en suministro_list
        suministros = Suministro.objects.select_related(
            'sst', 'distrito', 'estado_suministro'
        ).all()
        
        if sst_filter:
            suministros = suministros.filter(sst__sst__icontains=sst_filter)
        
        if distrito_filter:
            suministros = suministros.filter(distrito__nombre_distrito=distrito_filter)
        
        if estado_filter:
            suministros = suministros.filter(
                estado_suministro__estado_suministro=estado_filter
            )
        
        if search:
            suministros = suministros.filter(
                Q(suministro__icontains=search) |
                Q(medidor__icontains=search) |
                Q(direccion__icontains=search)
            )
        
        # Crear DataFrame
        data = []
        for s in suministros:
            data.append({
                'ITEM': s.item,
                'SST': s.sst.sst if s.sst else '',
                'SUMINISTRO': s.suministro,
                'TIPO': s.tipo_suministro,
                'MEDIDOR': s.medidor or '',
                'MARCA_MEDIDOR': s.marca_medidor or '',
                'FASE': s.fase or '',
                'POTENCIA': s.potencia or '',
                'ESTADO': s.estado_suministro.estado_suministro if s.estado_suministro else '',
                'DIRECCIÓN': s.direccion,
                'DISTRITO': s.distrito.nombre_distrito if s.distrito else '',
                'CONTACTO': s.contacto or '',
                'TELÉFONO': s.telefono or '',
                'LATITUD': s.latitud or '',
                'LONGITUD': s.longitud or '',
                'FECHA_PRIMER_ENVIO': s.fecha_primer_envio.strftime('%d/%m/%Y') if s.fecha_primer_envio else '',
                'FECHA_PROGRAMADA': s.fecha_programada.strftime('%d/%m/%Y') if s.fecha_programada else '',
                'HORA_INICIO_PROG': s.hora_inicio_programada.strftime('%H:%M') if s.hora_inicio_programada else '',
                'HORA_FIN_PROG': s.hora_fin_programada.strftime('%H:%M') if s.hora_fin_programada else '',
                'FECHA_EJECUCIÓN': s.fecha_ejecucion.strftime('%d/%m/%Y') if s.fecha_ejecucion else '',
                'EJECUTADO_POR': s.ejecutado_por or '',
                'OBSERVACIÓN_CONTRATISTA': s.observacion_contratista or '',
                'MOTIVO_ADICIONAL': s.motivo_adicional or '',
                'FECHA_IDENTIFICACIÓN': s.fecha_identificacion.strftime('%d/%m/%Y') if s.fecha_identificacion else '',
                'SOLICITADO_POR': s.solicitado_por or '',
            })
        
        df = pd.DataFrame(data)
        
        # Crear el archivo Excel en memoria
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Suministros', index=False)
            
            # Autoajustar columnas
            worksheet = writer.sheets['Suministros']
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        output.seek(0)
        
        # Crear respuesta HTTP
        filename = f"suministros_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
        
    except Exception as e:
        messages.error(request, f'Error al generar Excel: {str(e)}')
        return redirect('suministro_list')        
    
    
    
    
# management/commands/sincronizar_montos_sst.py
from django.core.management.base import BaseCommand
from gestion.models import SST
from decimal import Decimal

class Command(BaseCommand):
    help = 'Sincroniza los montos de todas las SSTs con la suma de sus suministros'
    
    def handle(self, *args, **kwargs):
        ssts = SST.objects.all()
        total_actualizado = 0
        
        for sst in ssts:
            monto_anterior = sst.monto_proyectado
            monto_nuevo = sst.actualizar_monto_total()
            
            if monto_anterior != monto_nuevo:
                self.stdout.write(
                    self.style.SUCCESS(
                        f'SST {sst.sst}: {monto_anterior} → {monto_nuevo}'
                    )
                )
                total_actualizado += 1
        
        self.stdout.write(
            self.style.SUCCESS(
                f'\n✅ {total_actualizado} SSTs actualizadas de {ssts.count()} totales'
            )
        )    
        
        
        
@login_required
def descargar_excel_modulo(request):
    pass

from django.db.models import Count, Sum, Q
from datetime import datetime, timedelta
from django.utils import timezone
from decimal import Decimal

@login_required
def dashboard(request):
    """Dashboard simple de SST"""
    
    hoy = timezone.now()
    inicio_mes = hoy.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    
    # KPIs principales
    sst_totales = SST.objects.count()
    
    # Estados de SST (EXACTAMENTE como mencionaste)
    sst_ejecutadas = SST.objects.filter(
        estado_sst__estado='EJECUTADO'
    ).count()
    
    sst_en_ejecucion = SST.objects.filter(
        estado_sst__estado='EN EJECUCIÓN'
    ).count()
    
    sst_admisibles = SST.objects.filter(
        estado_sst__estado='ADMISIBLE'
    ).count()
    
    # Otros estados (si existen)
    sst_otros = sst_totales - (sst_ejecutadas + sst_en_ejecucion + sst_admisibles)
    
    # Porcentajes
    if sst_totales > 0:
        porcentaje_ejecutadas = round((sst_ejecutadas / sst_totales) * 100, 1)
        porcentaje_en_ejecucion = round((sst_en_ejecucion / sst_totales) * 100, 1)
        porcentaje_admisibles = round((sst_admisibles / sst_totales) * 100, 1)
        porcentaje_otros = round((sst_otros / sst_totales) * 100, 1)
    else:
        porcentaje_ejecutadas = porcentaje_en_ejecucion = porcentaje_admisibles = porcentaje_otros = 0
    
    # Suministros (EXACTAMENTE como mencionaste)
    total_suministros = Suministro.objects.count()
    
    suministros_ejecutados = Suministro.objects.filter(
        estado_suministro__estado_suministro='EJECUTADO'
    ).count()
    
    suministros_asignados = Suministro.objects.filter(
        estado_suministro__estado_suministro='ASIGNADO'
    ).count()
    
    suministros_devueltos = Suministro.objects.filter(
        estado_suministro__estado_suministro='DEVUELTO'
    ).count()
    
    # Montos
    monto_total_proyectado = SST.objects.aggregate(
        total=Sum('monto_proyectado')
    )['total'] or Decimal('0.00')
    
    monto_total_real = SST.objects.aggregate(
        total=Sum('monto_real')
    )['total'] or Decimal('0.00')
    
    # Eficiencia financiera
    if monto_total_proyectado > 0:
        eficiencia_financiera = round((monto_total_real / monto_total_proyectado) * 100, 1)
    else:
        eficiencia_financiera = 0
    
    # Este mes
    sst_este_mes = SST.objects.filter(
        created_at__gte=inicio_mes
    ).count()
    
    sst_ejecutadas_mes = SST.objects.filter(
        created_at__gte=inicio_mes,
        estado_sst__estado='EJECUTADO'
    ).count()
    
    monto_mes = SST.objects.filter(
        created_at__gte=inicio_mes
    ).aggregate(
        total=Sum('monto_proyectado')
    )['total'] or Decimal('0.00')
    
    suministros_mes = Suministro.objects.filter(
        created_at__gte=inicio_mes
    ).count()
    
    # Últimas SST
    ultimas_sst = SST.objects.select_related('distrito', 'estado_sst').order_by('-created_at')[:10]
    
    # SST en ejecución con días transcurridos
    sst_en_ejecucion_list = SST.objects.filter(
        estado_sst__estado='EN EJECUCIÓN'
    ).select_related('distrito').order_by('-created_at')[:10]
    
    # Calcular días transcurridos
    for sst in sst_en_ejecucion_list:
        sst.dias_transcurridos = (timezone.now().date() - sst.created_at.date()).days
    
    context = {
        # SST
        'sst_totales': sst_totales,
        'sst_ejecutadas': sst_ejecutadas,
        'sst_en_ejecucion': sst_en_ejecucion,
        'sst_admisibles': sst_admisibles,
        'sst_otros': sst_otros,
        
        # Porcentajes SST
        'porcentaje_ejecutadas': porcentaje_ejecutadas,
        'porcentaje_en_ejecucion': porcentaje_en_ejecucion,
        'porcentaje_admisibles': porcentaje_admisibles,
        'porcentaje_otros': porcentaje_otros,
        
        # Suministros
        'total_suministros': total_suministros,
        'suministros_ejecutados': suministros_ejecutados,
        'suministros_asignados': suministros_asignados,
        'suministros_devueltos': suministros_devueltos,
        
        # Financiero
        'monto_total_proyectado': monto_total_proyectado,
        'monto_total_real': monto_total_real,
        'eficiencia_financiera': eficiencia_financiera,
        
        # Este mes
        'sst_este_mes': sst_este_mes,
        'sst_ejecutadas_mes': sst_ejecutadas_mes,
        'monto_mes': monto_mes,
        'suministros_mes': suministros_mes,
        
        # Listados
        'ultimas_sst': ultimas_sst,
        'sst_en_ejecucion_list': sst_en_ejecucion_list,
    }
    
    return render(request, 'gestion/dashboard.html', context)







# sst_app/views.


# sst_app/views.py

"""
from django.shortcuts import render
from django.db.models import Sum

def reporte_productividad(request):
  
    Vista para mostrar el reporte de productividad por ejecutor
    Maneja correctamente valores None y tablas vacías
   
    try:
        # Obtener suministros con ejecutor
        suministros_con_ejecutor = Suministro.objects.exclude(
            ejecutado_por__isnull=True
        ).exclude(
            ejecutado_por=''
        )

        # Agrupar por ejecutor y sumar montos
        resultados = suministros_con_ejecutor.values('ejecutado_por').annotate(
            total_monto=Sum('monto')
        ).order_by('ejecutado_por')

        # Calcular total general
        total_general = sum(item['total_monto'] or 0 for item in resultados)

        # Calcular promedio por ejecutor
        promedio_por_ejecutor = 0
        if len(resultados) > 0:
            promedio_por_ejecutor = total_general / len(resultados)

    except Exception as e:
        print(f"[ERROR] reporte_productividad: {e}")
        resultados = []
        total_general = 0
        promedio_por_ejecutor = 0

    context = {
        'resultados': resultados,
        'total_general': total_general,
        'promedio_por_ejecutor': promedio_por_ejecutor,
    }
    
    return render(request, 'gestion/reporte_productividad.html', context)"""
    
   
   
from collections import defaultdict
from decimal import Decimal
from datetime import datetime



from collections import defaultdict
from decimal import Decimal
from django.shortcuts import render

def reporte_productividad(request):
    """
    Reporte en formato MATRIZ: fechas en filas, ejecutores en columnas.
    """
    # Obtener filtros opcionales
    fecha_inicio = request.GET.get('fecha_inicio')
    fecha_fin = request.GET.get('fecha_fin')
    
    # Filtrar suministros EJECUTADOS y DEVUELTOS
    suministros = Suministro.objects.filter(
        ejecutado_por__isnull=False,
        fecha_ejecucion__isnull=False,
        estado_suministro__estado_suministro__in=['EJECUTADO', 'DEVUELTO','ADMISIBLE']
    ).exclude(ejecutado_por='')
    
    # Aplicar filtros de fecha si existen
    if fecha_inicio:
        suministros = suministros.filter(fecha_ejecucion__gte=fecha_inicio)
    if fecha_fin:
        suministros = suministros.filter(fecha_ejecucion__lte=fecha_fin)
    
    suministros = suministros.select_related('estado_suministro').order_by('fecha_ejecucion', 'ejecutado_por')

    # Diccionarios para la matriz
    matriz = defaultdict(lambda: defaultdict(lambda: Decimal('0.00')))
    ejecutores_set = set()
    totales_por_ejecutor = defaultdict(lambda: Decimal('0.00'))
    totales_por_fecha = defaultdict(lambda: Decimal('0.00'))
    total_general = Decimal('0.00')

    # Meses en español
    meses = {
        1: 'enero', 2: 'febrero', 3: 'marzo', 4: 'abril',
        5: 'mayo', 6: 'junio', 7: 'julio', 8: 'agosto',
        9: 'septiembre', 10: 'octubre', 11: 'noviembre', 12: 'diciembre'
    }

    # Procesar suministros y construir matriz
    fechas_ordenadas = []
    fechas_formateadas = {}
    
    for s in suministros:
        fecha = s.fecha_ejecucion
        ejecutor = s.ejecutado_por.strip().title()
        monto = s.monto or Decimal('0.00')

        # Formatear fecha
        if fecha not in fechas_formateadas:
            dia = fecha.day
            mes = meses[fecha.month]
            año = fecha.year
            fecha_formateada = f"{dia:02d} de {mes} de {año}"
            fechas_formateadas[fecha] = fecha_formateada
            fechas_ordenadas.append(fecha)

        # Llenar matriz
        matriz[fecha][ejecutor] += monto
        ejecutores_set.add(ejecutor)
        totales_por_ejecutor[ejecutor] += monto
        totales_por_fecha[fecha] += monto
        total_general += monto

    # Ordenar ejecutores alfabéticamente
    ejecutores = sorted(ejecutores_set)
    
    # Ordenar fechas
    fechas_ordenadas = sorted(set(fechas_ordenadas))

    # Construir estructura final para el template
    reporte = []
    for fecha in fechas_ordenadas:
        fila = {
            'fecha': fechas_formateadas[fecha],
            'montos': [matriz[fecha][ejecutor] for ejecutor in ejecutores],
            'total': totales_por_fecha[fecha]
        }
        reporte.append(fila)

    context = {
        'reporte': reporte,
        'ejecutores': ejecutores,
        'totales_por_ejecutor': [totales_por_ejecutor[ejecutor] for ejecutor in ejecutores],
        'total_general': total_general,
        'fecha_inicio': fecha_inicio,
        'fecha_fin': fecha_fin,
    }

    return render(request, 'gestion/reporte_productividad.html', context)










from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date

@login_required
def descargar_plantilla_importacion(request):
    """
    Genera una plantilla Excel con ejemplos para importar suministros
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Plantilla Importación"
    
    # Estilos
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    example_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ['SST', 'Suministro', 'Estado', 'Monto', 'Ejecutado por', 'Fecha de ejecucion']
    ws.append(headers)
    
    # Aplicar estilos a headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Ejemplos de datos
    ejemplos = [
        ['SST-001', 'SST-001-001', 'Ejecutado', '150.00', 'Juan Pérez', '15/01/2026'],
        ['SST-001', 'SST-001-002', 'Ejecutado', '200.50', 'María García', '16/01/2026'],
        ['SST-002', 'SST-002-001', 'Pendiente', '', '', ''],
        ['SST-002', 'SST-002-AD001', 'Asignado', '180.00', 'Carlos López', '17/01/2026'],
    ]
    
    for row_num, row_data in enumerate(ejemplos, 2):
        ws.append(row_data)
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.fill = example_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Ajustar anchos de columna
    column_widths = [15, 20, 15, 12, 20, 18]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width
    
    # Agregar hoja de instrucciones
    ws_instrucciones = wb.create_sheet("Instrucciones")
    
    instrucciones = [
        ["INSTRUCCIONES PARA IMPORTAR SUMINISTROS", ""],
        ["", ""],
        ["Columnas Obligatorias:", ""],
        ["✓ SST", "Código de la SST (debe existir en el sistema)"],
        ["✓ Suministro", "Número de suministro (debe existir en el sistema)"],
        ["✓ Estado", "Estado del suministro (Ejecutado, Pendiente, Asignado, etc.)"],
        ["", ""],
        ["Columnas Opcionales:", ""],
        ["○ Monto", "Monto en soles (número con hasta 2 decimales)"],
        ["○ Ejecutado por", "Nombre del técnico que ejecutó el trabajo"],
        ["○ Fecha de ejecucion", "Formato: DD/MM/YYYY"],
        ["", ""],
        ["NOTAS IMPORTANTES:", ""],
        ["1.", "Los suministros DEBEN existir previamente en el sistema"],
        ["2.", "Solo se actualizarán los campos incluidos en el Excel"],
        ["3.", "Los campos vacíos NO modificarán los datos existentes"],
        ["4.", "El monto debe ser un número positivo"],
        ["5.", "La fecha debe estar en formato DD/MM/YYYY"],
        ["6.", "El estado debe coincidir exactamente con los estados del sistema"],
        ["", ""],
        ["Estados válidos:", "Consultar en el sistema los estados disponibles"],
    ]
    
    for row_num, (col1, col2) in enumerate(instrucciones, 1):
        ws_instrucciones.cell(row=row_num, column=1, value=col1)
        ws_instrucciones.cell(row=row_num, column=2, value=col2)
        
        # Estilo para títulos
        if row_num == 1:
            ws_instrucciones.cell(row=row_num, column=1).font = Font(bold=True, size=14, color="4472C4")
        elif col1 in ["Columnas Obligatorias:", "Columnas Opcionales:", "NOTAS IMPORTANTES:", "Estados válidos:"]:
            ws_instrucciones.cell(row=row_num, column=1).font = Font(bold=True, size=11)
    
    ws_instrucciones.column_dimensions['A'].width = 25
    ws_instrucciones.column_dimensions['B'].width = 60
    
    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=plantilla_importacion_suministros_{date.today()}.xlsx'
    
    wb.save(response)
    return response



from django.contrib import messages
from django.shortcuts import redirect
from django.shortcuts import render
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.contrib.auth.decorators import login_required
from django.db import transaction
from openpyxl import load_workbook
from datetime import datetime
from decimal import Decimal, InvalidOperation
import traceback
from .models import Suministro, SST, EstadoSuministro

@login_required
@require_http_methods(["POST"])
def importar_excel_suministros(request):
    """
    Vista para importar y actualizar suministros desde un archivo Excel
    
    Columnas esperadas:
    - SST: Código de la SST
    - Suministro: Número de suministro
    - Estado: Estado del suministro
    - Monto: Monto en soles (opcional)
    - Ejecutado por: Nombre del ejecutor (opcional)
    - Fecha de ejecucion: Fecha en formato DD/MM/YYYY (opcional)
    """
    
    if 'archivo' not in request.FILES:
        return JsonResponse({
            'success': False,
            'message': 'No se ha seleccionado ningún archivo'
        })
    
    archivo = request.FILES['archivo']
    
    # Validar extensión del archivo
    if not archivo.name.endswith(('.xlsx', '.xls')):
        return JsonResponse({
            'success': False,
            'message': 'El archivo debe ser un Excel (.xlsx o .xls)'
        })
    
    try:
        # Cargar el archivo Excel
        wb = load_workbook(archivo, data_only=True)
        ws = wb.active
        
        # Validar que tenga al menos las columnas mínimas
        headers = [cell.value for cell in ws[1]]
        required_columns = ['SST', 'Suministro', 'Estado']
        
        # Normalizar headers (eliminar espacios y convertir a minúsculas)
        headers_normalized = {h.strip().lower(): idx for idx, h in enumerate(headers) if h}
        
        # Verificar columnas requeridas
        missing_columns = []
        for col in required_columns:
            if col.lower() not in headers_normalized:
                missing_columns.append(col)
        
        if missing_columns:
            return JsonResponse({
                'success': False,
                'message': f'Faltan columnas requeridas: {", ".join(missing_columns)}'
            })
        
        # Índices de columnas (base 0)
        col_sst = headers_normalized.get('sst')
        col_suministro = headers_normalized.get('suministro')
        col_estado = headers_normalized.get('estado')
        col_monto = headers_normalized.get('monto')
        col_ejecutado_por = headers_normalized.get('ejecutado por') or headers_normalized.get('ejecutado_por')
        col_fecha = headers_normalized.get('fecha de ejecucion') or headers_normalized.get('fecha_ejecucion') or headers_normalized.get('fecha de ejecución')
        
        # Contadores para el reporte
        actualizados = 0
        errores = []
        filas_procesadas = 0
        ssts_afectadas = set()  # ✅ Para actualizar SST solo una vez al final
        
        # ✅ USAR TRANSACCIÓN para asegurar atomicidad
        with transaction.atomic():
            # Procesar cada fila (saltando el encabezado)
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    # Saltar filas vacías
                    if not any(row):
                        continue
                    
                    filas_procesadas += 1
                    
                    # Extraer valores
                    sst_code = str(row[col_sst]).strip() if row[col_sst] else None
                    suministro_code = str(row[col_suministro]).strip() if row[col_suministro] else None
                    estado_nombre = str(row[col_estado]).strip() if row[col_estado] else None
                    
                    # Validaciones básicas
                    if not sst_code or not suministro_code or not estado_nombre:
                        errores.append({
                            'fila': row_num,
                            'error': 'SST, Suministro y Estado son obligatorios'
                        })
                        continue
                    
                    # Buscar la SST
                    try:
                        sst = SST.objects.get(sst=sst_code)
                        ssts_afectadas.add(sst.id)  # ✅ Registrar SST afectada
                    except SST.DoesNotExist:
                        errores.append({
                            'fila': row_num,
                            'sst': sst_code,
                            'error': f'SST "{sst_code}" no existe en el sistema'
                        })
                        continue
                    
                    # Buscar el suministro
                    try:
                        # ✅ OPTIMIZACIÓN: select_related para evitar consultas extra
                        suministro = Suministro.objects.select_related('sst', 'estado_suministro').get(
                            sst=sst,
                            suministro=suministro_code
                        )
                    except Suministro.DoesNotExist:
                        errores.append({
                            'fila': row_num,
                            'sst': sst_code,
                            'suministro': suministro_code,
                            'error': f'Suministro "{suministro_code}" no existe en SST "{sst_code}"'
                        })
                        continue
                    
                    # Buscar el estado
                    try:
                        estado = EstadoSuministro.objects.get(
                            estado_suministro__iexact=estado_nombre
                        )
                    except EstadoSuministro.DoesNotExist:
                        errores.append({
                            'fila': row_num,
                            'sst': sst_code,
                            'suministro': suministro_code,
                            'error': f'Estado "{estado_nombre}" no existe. Estados válidos: {", ".join(EstadoSuministro.objects.values_list("estado_suministro", flat=True))}'
                        })
                        continue
                    
                    # Actualizar el suministro
                    suministro.estado_suministro = estado
                    
                    # Procesar monto (opcional)
                    if col_monto is not None and row[col_monto]:
                        try:
                            monto_value = str(row[col_monto]).strip().replace(',', '.')
                            monto = Decimal(monto_value)
                            if monto < 0:
                                errores.append({
                                    'fila': row_num,
                                    'sst': sst_code,
                                    'suministro': suministro_code,
                                    'error': 'El monto no puede ser negativo',
                                    'advertencia': True
                                })
                            else:
                                suministro.monto = monto
                        except (InvalidOperation, ValueError):
                            errores.append({
                                'fila': row_num,
                                'sst': sst_code,
                                'suministro': suministro_code,
                                'error': f'Monto inválido: "{row[col_monto]}"',
                                'advertencia': True
                            })
                    
                    # Procesar ejecutado por (opcional)
                    if col_ejecutado_por is not None and row[col_ejecutado_por]:
                        suministro.ejecutado_por = str(row[col_ejecutado_por]).strip()
                    
                    # Procesar fecha de ejecución (opcional)
                    if col_fecha is not None and row[col_fecha]:
                        fecha_value = row[col_fecha]
                        try:
                            if isinstance(fecha_value, datetime):
                                suministro.fecha_ejecucion = fecha_value.date()
                            elif isinstance(fecha_value, str):
                                # Intentar diferentes formatos
                                for fmt in ['%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y']:
                                    try:
                                        suministro.fecha_ejecucion = datetime.strptime(fecha_value.strip(), fmt).date()
                                        break
                                    except ValueError:
                                        continue
                                else:
                                    errores.append({
                                        'fila': row_num,
                                        'sst': sst_code,
                                        'suministro': suministro_code,
                                        'error': f'Fecha inválida: "{fecha_value}". Use formato DD/MM/YYYY',
                                        'advertencia': True
                                    })
                        except Exception as e:
                            errores.append({
                                'fila': row_num,
                                'sst': sst_code,
                                'suministro': suministro_code,
                                'error': f'Error procesando fecha: {str(e)}',
                                'advertencia': True
                            })
                    
                    # ✅ GUARDAR SIN DISPARAR SEÑALES DE SST
                    # Usamos update_fields para evitar que se ejecute todo el save()
                    campos_actualizar = ['estado_suministro']
                    
                    if col_monto is not None and row[col_monto]:
                        campos_actualizar.append('monto')
                    if col_ejecutado_por is not None and row[col_ejecutado_por]:
                        campos_actualizar.append('ejecutado_por')
                    if col_fecha is not None and row[col_fecha]:
                        campos_actualizar.append('fecha_ejecucion')
                    
                    suministro.save(update_fields=campos_actualizar)
                    actualizados += 1
                    
                except Exception as e:
                    errores.append({
                        'fila': row_num,
                        'error': f'Error inesperado: {str(e)}'
                    })
                    continue
            
            # ✅ ACTUALIZAR SST SOLO UNA VEZ AL FINAL
            # Esto es mucho más eficiente que hacerlo en cada save()
            for sst_id in ssts_afectadas:
                try:
                    sst = SST.objects.get(id=sst_id)
                    sst.actualizar_monto_total()
                    sst.actualizar_estado_segun_suministros()
                except Exception as e:
                    print(f"⚠️ Error actualizando SST {sst_id}: {e}")
        
        # Preparar respuesta
        mensaje = f'Proceso completado: {actualizados} suministros actualizados de {filas_procesadas} filas procesadas'
        
        if errores:
            mensaje += f'. Se encontraron {len(errores)} errores o advertencias.'
            
        messages.success(request, mensaje)
        return redirect('suministro_list')
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error procesando el archivo: {str(e)}',
            'traceback': traceback.format_exc()
        })
        
 
import json
from django.shortcuts import render
from django.db.models import Count
from .models import Suministro

def mapa_suministros(request):
    # Obtener suministros con prefetch para optimizar
    qs = Suministro.objects.filter(
        estado_suministro__estado_suministro='ASIGNADO',
        latitud__isnull=False,
        longitud__isnull=False
    ).exclude(
        latitud=0,
        longitud=0
    ).select_related(
        'sst', 
        'estado_suministro'
    ).values(
        'suministro',
        'direccion',
        'latitud',
        'longitud',
        'sst__sst',
        'sst__id',  # Para agrupar por SST
        'estado_suministro__estado_suministro',
        'estado_suministro__color',
        'medidor',
        'potencia',
        'contacto',
        'telefono'
    )

    # Generar colores únicos por SST
    ssts_unicos = list(set([s['sst__id'] for s in qs]))
    colores_sst = generar_colores_distintos(len(ssts_unicos))
    mapa_colores = {sst_id: color for sst_id, color in zip(ssts_unicos, colores_sst)}

    suministros = []
    for s in qs:
        try:
            lat = float(s['latitud'])
            lng = float(s['longitud'])
            
            if -90 <= lat <= 90 and -180 <= lng <= 180:
                suministros.append({
                    'suministro': s['suministro'],
                    'direccion': s['direccion'] or '',
                    'latitud': lat,
                    'longitud': lng,
                    'sst': s['sst__sst'] or '',
                    'sst_id': s['sst__id'],
                    'color': mapa_colores.get(s['sst__id'], '#3388ff'),
                    'estado': s['estado_suministro__estado_suministro'],
                    'estado_color': s['estado_suministro__color'] or '#3388ff',
                    'medidor': s['medidor'] or 'N/A',
                    'potencia': s['potencia'] or 'N/A',
                    'contacto': s['contacto'] or 'N/A',
                    'telefono': s['telefono'] or 'N/A',
                })
        except (ValueError, TypeError):
            continue

    # Agrupar por SST para la leyenda
    ssts_info = {}
    for s in suministros:
        sst_id = s['sst_id']
        if sst_id not in ssts_info:
            ssts_info[sst_id] = {
                'nombre': s['sst'],
                'color': s['color'],
                'cantidad': 0
            }
        ssts_info[sst_id]['cantidad'] += 1

    return render(request, "gestion/mapa_suministros.html", {
        "suministros": json.dumps(suministros),
        "ssts_info": json.dumps(list(ssts_info.values())),
        "total": len(suministros)
    })


def generar_colores_distintos(n):
    """Genera n colores visualmente distintos usando HSL"""
    import colorsys
    colores = []
    for i in range(n):
        hue = i / n
        # Saturación y luminosidad fijas para colores vibrantes
        rgb = colorsys.hsv_to_rgb(hue, 0.8, 0.9)
        hex_color = '#{:02x}{:02x}{:02x}'.format(
            int(rgb[0] * 255),
            int(rgb[1] * 255),
            int(rgb[2] * 255)
        )
        colores.append(hex_color)
    return colores       